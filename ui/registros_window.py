"""
Janela de Visualização de Registros
Interface para visualizar histórico de entradas e saídas
"""

import os
import csv
from datetime import date, datetime
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QFrame, QTableWidget, QTableWidgetItem,
    QHeaderView, QDateEdit, QComboBox, QMessageBox, QFileDialog, QMenu
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor

from database.models import Database
from core.config import get_config

# Importações opcionais para exportação
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class RegistrosWindow(QDialog):
    """Janela para visualização de registros"""

    def __init__(self, db: Database, parent=None):
        super().__init__(parent)

        self.db = db

        # Configura interface
        self._setup_ui()

        # Carrega registros do dia
        self._carregar_registros()

    def _setup_ui(self):
        """Configura a interface gráfica"""
        self.setWindowTitle("Registros de Acesso")
        self.setMinimumSize(900, 600)
        self.setStyleSheet(self._get_stylesheet())

        # Layout principal
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # ==================== CABEÇALHO ====================
        header = QFrame()
        header.setObjectName("header")
        header_layout = QHBoxLayout(header)

        titulo = QLabel("REGISTROS DE ACESSO")
        titulo.setObjectName("titulo")
        header_layout.addWidget(titulo)

        header_layout.addStretch()

        # Filtros
        filtros_layout = QHBoxLayout()
        filtros_layout.setSpacing(10)

        # Data
        filtros_layout.addWidget(QLabel("Data:"))
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("dd/MM/yyyy")
        self.date_edit.dateChanged.connect(self._carregar_registros)
        filtros_layout.addWidget(self.date_edit)

        # Tipo
        filtros_layout.addWidget(QLabel("Tipo:"))
        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems(["Todos", "Entrada", "Saída"])
        self.tipo_combo.currentIndexChanged.connect(self._aplicar_filtro)
        filtros_layout.addWidget(self.tipo_combo)

        header_layout.addLayout(filtros_layout)

        main_layout.addWidget(header)

        # ==================== RESUMO ====================
        resumo_frame = QFrame()
        resumo_frame.setObjectName("resumoFrame")
        resumo_layout = QHBoxLayout(resumo_frame)

        # Total de registros
        self.total_label = QLabel("Total: 0")
        self.total_label.setObjectName("resumoItem")
        resumo_layout.addWidget(self.total_label)

        resumo_layout.addStretch()

        # Entradas
        self.entradas_label = QLabel("Entradas: 0")
        self.entradas_label.setObjectName("resumoEntrada")
        resumo_layout.addWidget(self.entradas_label)

        # Saídas
        self.saidas_label = QLabel("Saídas: 0")
        self.saidas_label.setObjectName("resumoSaida")
        resumo_layout.addWidget(self.saidas_label)

        main_layout.addWidget(resumo_frame)

        # ==================== TABELA ====================
        self.tabela = QTableWidget()
        self.tabela.setObjectName("tabelaRegistros")
        self.tabela.setColumnCount(7)
        self.tabela.setHorizontalHeaderLabels([
            "Horário", "Nome", "Matrícula", "Turma", "Tipo", "Confiança", "Modo"
        ])

        # Configura colunas
        header = self.tabela.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # Horário
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # Nome
        header.setSectionResizeMode(2, QHeaderView.Fixed)  # Matrícula
        header.setSectionResizeMode(3, QHeaderView.Fixed)  # Turma
        header.setSectionResizeMode(4, QHeaderView.Fixed)  # Tipo
        header.setSectionResizeMode(5, QHeaderView.Fixed)  # Confiança
        header.setSectionResizeMode(6, QHeaderView.Fixed)  # Modo

        self.tabela.setColumnWidth(0, 80)
        self.tabela.setColumnWidth(2, 100)
        self.tabela.setColumnWidth(3, 70)
        self.tabela.setColumnWidth(4, 80)
        self.tabela.setColumnWidth(5, 90)
        self.tabela.setColumnWidth(6, 70)

        # Configurações da tabela
        self.tabela.setAlternatingRowColors(True)
        self.tabela.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabela.setSelectionMode(QTableWidget.SingleSelection)
        self.tabela.verticalHeader().setVisible(False)
        self.tabela.setEditTriggers(QTableWidget.NoEditTriggers)

        main_layout.addWidget(self.tabela)

        # ==================== BOTÕES ====================
        botoes_layout = QHBoxLayout()

        # Atualizar
        btn_atualizar = QPushButton("ATUALIZAR")
        btn_atualizar.setObjectName("btnAtualizar")
        btn_atualizar.clicked.connect(self._carregar_registros)
        botoes_layout.addWidget(btn_atualizar)

        botoes_layout.addStretch()

        # Exportar com menu de opções
        self.btn_exportar = QPushButton("EXPORTAR ▼")
        self.btn_exportar.setObjectName("btnExportar")
        self.btn_exportar.clicked.connect(self._mostrar_menu_exportar)
        botoes_layout.addWidget(self.btn_exportar)

        # Fechar
        btn_fechar = QPushButton("FECHAR")
        btn_fechar.setObjectName("btnFechar")
        btn_fechar.clicked.connect(self.accept)
        botoes_layout.addWidget(btn_fechar)

        main_layout.addLayout(botoes_layout)

    def _carregar_registros(self):
        """Carrega os registros da data selecionada"""
        data_selecionada = self.date_edit.date().toPyDate()

        # Busca registros do banco
        self.registros = self.db.listar_registros_do_dia(data_selecionada)

        # Aplica filtro e atualiza tabela
        self._aplicar_filtro()

    def _aplicar_filtro(self):
        """Aplica filtro de tipo aos registros"""
        tipo_filtro = self.tipo_combo.currentText()

        # Filtra registros
        if tipo_filtro == "Todos":
            registros_filtrados = self.registros
        elif tipo_filtro == "Entrada":
            registros_filtrados = [r for r in self.registros if r.tipo == "entrada"]
        else:
            registros_filtrados = [r for r in self.registros if r.tipo == "saida"]

        # Atualiza tabela
        self._atualizar_tabela(registros_filtrados)

        # Atualiza resumo
        self._atualizar_resumo()

    def _atualizar_tabela(self, registros):
        """Atualiza a tabela com os registros"""
        self.tabela.setRowCount(len(registros))

        for row, registro in enumerate(registros):
            # Horário
            horario = registro.data_hora.strftime("%H:%M:%S") if registro.data_hora else "--:--:--"
            item_horario = QTableWidgetItem(horario)
            item_horario.setTextAlignment(Qt.AlignCenter)
            self.tabela.setItem(row, 0, item_horario)

            # Nome
            item_nome = QTableWidgetItem(registro.aluno_nome or "---")
            self.tabela.setItem(row, 1, item_nome)

            # Matrícula
            item_matricula = QTableWidgetItem(registro.aluno_matricula or "---")
            item_matricula.setTextAlignment(Qt.AlignCenter)
            self.tabela.setItem(row, 2, item_matricula)

            # Turma
            item_turma = QTableWidgetItem(registro.aluno_turma or "---")
            item_turma.setTextAlignment(Qt.AlignCenter)
            self.tabela.setItem(row, 3, item_turma)

            # Tipo
            tipo_texto = "ENTRADA" if registro.tipo == "entrada" else "SAÍDA"
            item_tipo = QTableWidgetItem(tipo_texto)
            item_tipo.setTextAlignment(Qt.AlignCenter)
            if registro.tipo == "entrada":
                item_tipo.setForeground(QColor("#27ae60"))
            else:
                item_tipo.setForeground(QColor("#e74c3c"))
            self.tabela.setItem(row, 4, item_tipo)

            # Confiança
            confianca_texto = f"{registro.confianca:.1f}%"
            item_confianca = QTableWidgetItem(confianca_texto)
            item_confianca.setTextAlignment(Qt.AlignCenter)
            self.tabela.setItem(row, 5, item_confianca)

            # Modo
            modo_texto = "Manual" if registro.manual else "Auto"
            item_modo = QTableWidgetItem(modo_texto)
            item_modo.setTextAlignment(Qt.AlignCenter)
            if registro.manual:
                item_modo.setForeground(QColor("#f39c12"))
            self.tabela.setItem(row, 6, item_modo)

    def _atualizar_resumo(self):
        """Atualiza o resumo de registros"""
        total = len(self.registros)
        entradas = len([r for r in self.registros if r.tipo == "entrada"])
        saidas = len([r for r in self.registros if r.tipo == "saida"])

        self.total_label.setText(f"Total: {total}")
        self.entradas_label.setText(f"Entradas: {entradas}")
        self.saidas_label.setText(f"Saídas: {saidas}")

    def _mostrar_menu_exportar(self):
        """Mostra menu com opções de exportação"""
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: #16213e;
                color: white;
                border: 1px solid #3a3a5c;
                padding: 5px;
            }
            QMenu::item {
                padding: 8px 25px;
            }
            QMenu::item:selected {
                background-color: #e94560;
            }
            QMenu::item:disabled {
                color: #666666;
            }
        """)

        # Opção CSV
        action_csv = menu.addAction("📄 Exportar CSV")
        action_csv.triggered.connect(self._exportar_csv)

        # Opção Excel
        action_excel = menu.addAction("📊 Exportar Excel (.xlsx)")
        if not OPENPYXL_AVAILABLE:
            action_excel.setEnabled(False)
            action_excel.setText("📊 Excel (instale openpyxl)")
        else:
            action_excel.triggered.connect(self._exportar_excel)

        # Opção PDF
        action_pdf = menu.addAction("📑 Exportar PDF")
        if not REPORTLAB_AVAILABLE:
            action_pdf.setEnabled(False)
            action_pdf.setText("📑 PDF (instale reportlab)")
        else:
            action_pdf.triggered.connect(self._exportar_pdf)

        # Mostra o menu abaixo do botão
        menu.exec_(self.btn_exportar.mapToGlobal(self.btn_exportar.rect().bottomLeft()))

    def _get_registros_para_exportar(self):
        """Retorna os registros filtrados para exportação"""
        tipo_filtro = self.tipo_combo.currentText()

        if tipo_filtro == "Todos":
            return self.registros
        elif tipo_filtro == "Entrada":
            return [r for r in self.registros if r.tipo == "entrada"]
        else:
            return [r for r in self.registros if r.tipo == "saida"]

    def _exportar_csv(self):
        """Exporta registros para arquivo CSV"""
        registros = self._get_registros_para_exportar()
        config = get_config()

        if not registros:
            QMessageBox.warning(self, "Aviso", "Não há registros para exportar.")
            return

        # Diálogo para salvar arquivo
        data_str = self.date_edit.date().toString("yyyy-MM-dd")
        nome_padrao = f"registros_{data_str}.csv"

        arquivo, _ = QFileDialog.getSaveFileName(
            self, "Salvar CSV", nome_padrao, "Arquivos CSV (*.csv)"
        )

        if not arquivo:
            return

        try:
            with open(arquivo, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')

                # Informações da escola
                writer.writerow([config.nome_completo_escola])
                writer.writerow([f"Registros de Acesso - {self.date_edit.date().toString('dd/MM/yyyy')}"])
                writer.writerow([])  # Linha em branco

                # Cabeçalho
                writer.writerow(["Data", "Horário", "Nome", "Matrícula", "Turma", "Tipo", "Confiança (%)", "Modo"])

                # Dados
                for reg in registros:
                    data = reg.data_hora.strftime("%d/%m/%Y") if reg.data_hora else ""
                    horario = reg.data_hora.strftime("%H:%M:%S") if reg.data_hora else ""
                    tipo = "Entrada" if reg.tipo == "entrada" else "Saída"
                    modo = "Manual" if reg.manual else "Automático"

                    writer.writerow([
                        data,
                        horario,
                        reg.aluno_nome or "",
                        reg.aluno_matricula or "",
                        reg.aluno_turma or "",
                        tipo,
                        f"{reg.confianca:.1f}",
                        modo
                    ])

                # Rodapé com créditos
                writer.writerow([])
                writer.writerow([f"Gerado por: Guardião Escolar v{config.config.versao}"])
                writer.writerow([config.creditos_desenvolvedor])

            QMessageBox.information(self, "Sucesso", f"Arquivo CSV exportado com sucesso!\n\n{arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar CSV:\n{str(e)}")

    def _exportar_excel(self):
        """Exporta registros para arquivo Excel"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Aviso", "Biblioteca openpyxl não instalada.\n\nExecute: pip install openpyxl")
            return

        registros = self._get_registros_para_exportar()
        config = get_config()

        if not registros:
            QMessageBox.warning(self, "Aviso", "Não há registros para exportar.")
            return

        # Diálogo para salvar arquivo
        data_str = self.date_edit.date().toString("yyyy-MM-dd")
        nome_padrao = f"registros_{data_str}.xlsx"

        arquivo, _ = QFileDialog.getSaveFileName(
            self, "Salvar Excel", nome_padrao, "Arquivos Excel (*.xlsx)"
        )

        if not arquivo:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Registros"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            cell_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            entrada_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            saida_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

            # Nome da escola
            ws.merge_cells('A1:H1')
            ws['A1'] = config.nome_completo_escola
            ws['A1'].font = Font(bold=True, size=16, color="E94560")
            ws['A1'].alignment = Alignment(horizontal="center")

            # Título
            data_formatada = self.date_edit.date().toString("dd/MM/yyyy")
            ws.merge_cells('A2:H2')
            ws['A2'] = f"Registros de Acesso - {data_formatada}"
            ws['A2'].font = Font(bold=True, size=12)
            ws['A2'].alignment = Alignment(horizontal="center")

            # Cabeçalho
            headers = ["Data", "Horário", "Nome", "Matrícula", "Turma", "Tipo", "Confiança (%)", "Modo"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Dados
            for row_idx, reg in enumerate(registros, 5):
                data = reg.data_hora.strftime("%d/%m/%Y") if reg.data_hora else ""
                horario = reg.data_hora.strftime("%H:%M:%S") if reg.data_hora else ""
                tipo = "Entrada" if reg.tipo == "entrada" else "Saída"
                modo = "Manual" if reg.manual else "Automático"

                row_data = [
                    data,
                    horario,
                    reg.aluno_nome or "",
                    reg.aluno_matricula or "",
                    reg.aluno_turma or "",
                    tipo,
                    reg.confianca,
                    modo
                ]

                row_fill = entrada_fill if reg.tipo == "entrada" else saida_fill

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    cell.fill = row_fill

            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 14
            ws.column_dimensions['H'].width = 12

            # Resumo
            total = len(registros)
            entradas = len([r for r in registros if r.tipo == "entrada"])
            saidas = len([r for r in registros if r.tipo == "saida"])
            linha_resumo = len(registros) + 6

            ws.cell(row=linha_resumo, column=1, value="Resumo:").font = Font(bold=True)
            ws.cell(row=linha_resumo + 1, column=1, value=f"Total: {total}")
            ws.cell(row=linha_resumo + 2, column=1, value=f"Entradas: {entradas}")
            ws.cell(row=linha_resumo + 3, column=1, value=f"Saídas: {saidas}")

            # Créditos do desenvolvedor
            linha_creditos = linha_resumo + 5
            ws.merge_cells(f'A{linha_creditos}:H{linha_creditos}')
            ws.cell(row=linha_creditos, column=1, value=f"Gerado por: Guardião Escolar v{config.config.versao}")
            ws.cell(row=linha_creditos, column=1).font = Font(italic=True, size=9, color="666666")
            ws.cell(row=linha_creditos, column=1).alignment = Alignment(horizontal="center")

            ws.merge_cells(f'A{linha_creditos + 1}:H{linha_creditos + 1}')
            ws.cell(row=linha_creditos + 1, column=1, value=config.creditos_desenvolvedor)
            ws.cell(row=linha_creditos + 1, column=1).font = Font(italic=True, size=9, color="3498DB")
            ws.cell(row=linha_creditos + 1, column=1).alignment = Alignment(horizontal="center")

            wb.save(arquivo)
            QMessageBox.information(self, "Sucesso", f"Arquivo Excel exportado com sucesso!\n\n{arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar Excel:\n{str(e)}")

    def _exportar_pdf(self):
        """Exporta registros para arquivo PDF"""
        if not REPORTLAB_AVAILABLE:
            QMessageBox.warning(self, "Aviso", "Biblioteca reportlab não instalada.\n\nExecute: pip install reportlab")
            return

        registros = self._get_registros_para_exportar()
        config = get_config()

        if not registros:
            QMessageBox.warning(self, "Aviso", "Não há registros para exportar.")
            return

        # Diálogo para salvar arquivo
        data_str = self.date_edit.date().toString("yyyy-MM-dd")
        nome_padrao = f"registros_{data_str}.pdf"

        arquivo, _ = QFileDialog.getSaveFileName(
            self, "Salvar PDF", nome_padrao, "Arquivos PDF (*.pdf)"
        )

        if not arquivo:
            return

        try:
            # Cria documento PDF em landscape para caber mais colunas
            doc = SimpleDocTemplate(
                arquivo,
                pagesize=landscape(A4),
                rightMargin=15*mm,
                leftMargin=15*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            elements = []
            styles = getSampleStyleSheet()

            # Nome da escola
            escola_style = ParagraphStyle(
                'Escola',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.HexColor('#E94560'),
                alignment=1,  # Center
                spaceAfter=5
            )
            elements.append(Paragraph(config.nome_completo_escola, escola_style))

            # Título
            data_formatada = self.date_edit.date().toString("dd/MM/yyyy")
            titulo_style = ParagraphStyle(
                'Titulo',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#333333'),
                alignment=1,  # Center
                spaceAfter=20
            )
            elements.append(Paragraph(f"Registros de Acesso - {data_formatada}", titulo_style))

            # Resumo
            total = len(registros)
            entradas = len([r for r in registros if r.tipo == "entrada"])
            saidas = len([r for r in registros if r.tipo == "saida"])

            resumo_style = ParagraphStyle(
                'Resumo',
                parent=styles['Normal'],
                fontSize=11,
                alignment=1,
                spaceAfter=15
            )
            elements.append(Paragraph(
                f"<b>Total:</b> {total} registros | <b>Entradas:</b> {entradas} | <b>Saídas:</b> {saidas}",
                resumo_style
            ))
            elements.append(Spacer(1, 10))

            # Tabela de dados
            table_data = [["Horário", "Nome", "Matrícula", "Turma", "Tipo", "Confiança", "Modo"]]

            for reg in registros:
                horario = reg.data_hora.strftime("%H:%M:%S") if reg.data_hora else "--:--:--"
                tipo = "Entrada" if reg.tipo == "entrada" else "Saída"
                modo = "Manual" if reg.manual else "Auto"

                table_data.append([
                    horario,
                    reg.aluno_nome or "---",
                    reg.aluno_matricula or "---",
                    reg.aluno_turma or "---",
                    tipo,
                    f"{reg.confianca:.1f}%",
                    modo
                ])

            # Cria tabela com larguras proporcionais
            col_widths = [60, 180, 80, 60, 60, 60, 50]
            table = Table(table_data, colWidths=col_widths)

            # Estilo da tabela
            table_style = TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E94560')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),

                # Corpo
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Nome alinhado à esquerda
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),

                # Bordas
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#3a3a5c')),
            ])

            # Cores alternadas e cor por tipo
            for i, reg in enumerate(registros, 1):
                if reg.tipo == "entrada":
                    bg_color = colors.HexColor('#D5F5E3')
                else:
                    bg_color = colors.HexColor('#FADBD8')
                table_style.add('BACKGROUND', (0, i), (-1, i), bg_color)

            table.setStyle(table_style)
            elements.append(table)

            # Rodapé
            elements.append(Spacer(1, 20))
            rodape_style = ParagraphStyle(
                'Rodape',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=1
            )
            agora = datetime.now().strftime("%d/%m/%Y às %H:%M")
            elements.append(Paragraph(f"Relatório gerado em {agora}", rodape_style))

            # Créditos do desenvolvedor
            creditos_style = ParagraphStyle(
                'Creditos',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#3498db'),
                alignment=1,
                spaceBefore=10
            )
            elements.append(Paragraph(
                f"<b>Guardião Escolar v{config.config.versao}</b><br/>"
                f"{config.creditos_desenvolvedor}",
                creditos_style
            ))

            # Gera PDF
            doc.build(elements)
            QMessageBox.information(self, "Sucesso", f"Arquivo PDF exportado com sucesso!\n\n{arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar PDF:\n{str(e)}")

    def _get_stylesheet(self):
        """Retorna o stylesheet da janela"""
        return """
            QDialog {
                background-color: #1a1a2e;
            }

            QFrame#header {
                background-color: #16213e;
                border-radius: 10px;
                padding: 15px;
            }

            QLabel#titulo {
                color: #e94560;
                font-size: 20px;
                font-weight: bold;
            }

            QFrame#resumoFrame {
                background-color: #16213e;
                border-radius: 8px;
                padding: 10px 20px;
            }

            QLabel#resumoItem {
                color: #ffffff;
                font-size: 16px;
                font-weight: bold;
            }

            QLabel#resumoEntrada {
                color: #27ae60;
                font-size: 16px;
                font-weight: bold;
            }

            QLabel#resumoSaida {
                color: #e74c3c;
                font-size: 16px;
                font-weight: bold;
            }

            QLabel {
                color: #ffffff;
                font-size: 14px;
            }

            QDateEdit {
                background-color: #0f0f1a;
                color: white;
                border: 1px solid #3a3a5c;
                border-radius: 5px;
                padding: 8px;
                font-size: 14px;
                min-width: 120px;
            }

            QDateEdit:focus {
                border-color: #e94560;
            }

            QDateEdit::drop-down {
                border: none;
                padding-right: 10px;
            }

            QComboBox {
                background-color: #0f0f1a;
                color: white;
                border: 1px solid #3a3a5c;
                border-radius: 5px;
                padding: 8px;
                font-size: 14px;
                min-width: 100px;
            }

            QComboBox:focus {
                border-color: #e94560;
            }

            QComboBox::drop-down {
                border: none;
                padding-right: 10px;
            }

            QComboBox QAbstractItemView {
                background-color: #16213e;
                color: white;
                selection-background-color: #e94560;
            }

            QTableWidget {
                background-color: #16213e;
                color: #ffffff;
                border: 1px solid #3a3a5c;
                border-radius: 8px;
                gridline-color: #3a3a5c;
                font-size: 13px;
            }

            QTableWidget::item {
                padding: 10px;
            }

            QTableWidget::item:selected {
                background-color: #e94560;
            }

            QTableWidget::item:alternate {
                background-color: #1a1a2e;
            }

            QHeaderView::section {
                background-color: #0f0f1a;
                color: #e94560;
                font-weight: bold;
                padding: 12px;
                border: none;
                border-bottom: 2px solid #e94560;
            }

            QPushButton#btnAtualizar {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 12px 25px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }

            QPushButton#btnAtualizar:hover {
                background-color: #5dade2;
            }

            QPushButton#btnExportar {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 12px 25px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }

            QPushButton#btnExportar:hover {
                background-color: #2ecc71;
            }

            QPushButton#btnExportar:disabled {
                background-color: #1e5f3b;
            }

            QPushButton#btnFechar {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }

            QPushButton#btnFechar:hover {
                background-color: #ec7063;
            }
        """
